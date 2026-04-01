import math
from copy import copy
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from arch import arch_model
from openpyxl import load_workbook
from scipy.stats import chi2


PROJECT_DIR = Path("/Users/deepikanath/dnath796/Git/Financial_Models/P/Project _01")
TEMPLATE_PATH = Path("/Users/deepikanath/Downloads/Volatility Class Work (2).xlsx")
DATA_PATH = Path("/Users/deepikanath/dnath796/Git/Financial_Models/Data_center/OVXCLS.csv")
OUTPUT_PATH = PROJECT_DIR / "Volatility Class Work - OVX.xlsx"


def copy_style(source_cell, target_cell):
    target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def ensure_row_style(ws, target_row, template_row, start_col, end_col):
    for col_idx in range(start_col, end_col + 1):
        copy_style(ws.cell(template_row, col_idx), ws.cell(target_row, col_idx))
    ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height


def load_ovx_data():
    df = pd.read_csv(DATA_PATH, sep=None, engine="python")
    df.columns = df.columns.str.strip()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
    df = df.dropna(subset=["Date", "Value"]).sort_values("Date").reset_index(drop=True)
    return df


def fit_models(df):
    prices = df["Value"]
    returns = prices.pct_change()

    reg_df = pd.DataFrame({"return": returns, "lag1": returns.shift(1)}).dropna().reset_index(drop=True)
    X = sm.add_constant(reg_df["lag1"])
    ols = sm.OLS(reg_df["return"], X).fit()

    residuals = ols.resid.to_numpy()
    garch_fit = arch_model(
        residuals,
        mean="Zero",
        vol="GARCH",
        p=1,
        q=1,
        dist="normal",
        rescale=False,
    ).fit(disp="off")

    return returns, reg_df, ols, residuals, garch_fit


def build_table(df, returns, reg_df, ols, residuals, garch_fit):
    n_prices = len(df)
    data_start_row = 6
    last_data_row = data_start_row + n_prices - 1

    intercept = float(ols.params["const"])
    phi = float(ols.params["lag1"])
    omega = float(garch_fit.params["omega"])
    alpha = float(garch_fit.params["alpha[1]"])
    beta = float(garch_fit.params["beta[1]"])
    unconditional_variance = omega / (1 - alpha - beta)
    ewma_weight = 1 - 0.94

    dates = df["Date"].tolist()
    prices = df["Value"].tolist()
    return_values = returns.tolist()

    fitted_values = ols.fittedvalues.to_numpy()
    regression_residuals = residuals

    predicted_by_row = {}
    residual_by_row = {}
    ewma_var_by_row = {}
    garch_var_by_row = {}
    restricted_var_by_row = {}

    for idx in range(2, n_prices):
        row_number = data_start_row + idx
        reg_pos = idx - 2
        fitted = float(fitted_values[reg_pos])
        residual = float(regression_residuals[reg_pos])
        predicted_by_row[row_number] = fitted
        residual_by_row[row_number] = residual

        if row_number == 8:
            ewma_var = residual**2
            garch_var = unconditional_variance
            restricted_var = unconditional_variance
        else:
            prev_row = row_number - 1
            ewma_var = ewma_weight * (residual**2) + (1 - ewma_weight) * ewma_var_by_row[prev_row]
            prev_residual = residual_by_row[prev_row]
            prev_garch_var = garch_var_by_row[prev_row]
            garch_var = omega + alpha * (prev_residual**2) + beta * prev_garch_var
            restricted_var = omega

        ewma_var_by_row[row_number] = ewma_var
        garch_var_by_row[row_number] = garch_var
        restricted_var_by_row[row_number] = restricted_var

    stats = {
        "intercept": intercept,
        "phi": phi,
        "historical_vol": float(reg_df["return"].std(ddof=1)),
        "returns_mean": float(reg_df["return"].mean()),
        "ewma_weight": ewma_weight,
        "omega": omega,
        "alpha": alpha,
        "beta": beta,
        "alpha_plus_beta": alpha + beta,
        "unconditional_variance": unconditional_variance,
        "ln_2pi": math.log(2 * math.pi),
        "multiple_r": math.sqrt(max(ols.rsquared, 0.0)),
        "rsquared": float(ols.rsquared),
        "rsquared_adj": float(ols.rsquared_adj),
        "standard_error": float(math.sqrt(ols.mse_resid)),
        "observations": int(ols.nobs),
        "regression_ss": float(ols.ess),
        "residual_ss": float(ols.ssr),
        "total_ss": float(ols.centered_tss),
        "regression_ms": float(ols.ess / max(int(ols.df_model), 1)),
        "residual_ms": float(ols.mse_resid),
        "f_stat": float(ols.fvalue),
        "f_pvalue": float(ols.f_pvalue),
        "df_model": int(ols.df_model),
        "df_resid": int(ols.df_resid),
        "df_total": int(ols.df_model + ols.df_resid),
        "conf_int": ols.conf_int(),
    }

    log_like_full = -0.5 * (
        sum(garch_var_by_row.values())
        + sum((residual_by_row[row] ** 2) / garch_var_by_row[row] for row in garch_var_by_row)
        + len(garch_var_by_row) * stats["ln_2pi"]
    )
    log_like_restricted = -0.5 * (
        sum(restricted_var_by_row.values())
        + sum((residual_by_row[row] ** 2) / restricted_var_by_row[row] for row in restricted_var_by_row)
        + len(restricted_var_by_row) * stats["ln_2pi"]
    )
    stats["llf"] = float(log_like_full)
    stats["restricted_llf"] = float(log_like_restricted)
    stats["lr_stat"] = float(-2 * (log_like_restricted - log_like_full))
    stats["chi_sq_95"] = float(chi2.ppf(0.95, 2))
    stats["std_errs"] = [float(ols.bse.iloc[0]), float(ols.bse.iloc[1])]
    stats["t_stats"] = [float(ols.tvalues.iloc[0]), float(ols.tvalues.iloc[1])]
    stats["p_values"] = [float(ols.pvalues.iloc[0]), float(ols.pvalues.iloc[1])]

    return {
        "dates": dates,
        "prices": prices,
        "returns": return_values,
        "predicted_by_row": predicted_by_row,
        "residual_by_row": residual_by_row,
        "ewma_var_by_row": ewma_var_by_row,
        "garch_var_by_row": garch_var_by_row,
        "restricted_var_by_row": restricted_var_by_row,
        "stats": stats,
        "last_data_row": last_data_row,
    }


def write_workbook(table):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]
    original_sheet_name = ws.title
    ws.title = "OVX"

    data_template_row = 3825
    residual_template_row = 3843
    data_start_row = 6
    residual_start_row = 26
    last_data_row = table["last_data_row"]
    last_residual_row = residual_start_row + table["stats"]["observations"] - 1

    for row in range(ws.max_row + 1, max(last_data_row, last_residual_row) + 1):
        if row <= last_data_row:
            ensure_row_style(ws, row, data_template_row, 1, 14)
        if row <= last_residual_row:
            ensure_row_style(ws, row, residual_template_row, 26, 28)

    ws["B5"] = "OVXCLS - Index Value"

    for idx, (date_value, price_value) in enumerate(zip(table["dates"], table["prices"]), start=data_start_row):
        ws.cell(idx, 1, date_value)
        ws.cell(idx, 2, float(price_value))

        rel_idx = idx - data_start_row
        return_value = table["returns"][rel_idx]
        ws.cell(idx, 3, None if pd.isna(return_value) else float(return_value))

        if idx >= 8:
            lag_value = table["returns"][rel_idx - 1]
            fitted = table["predicted_by_row"][idx]
            residual = table["residual_by_row"][idx]
            ewma_var = table["ewma_var_by_row"][idx]
            garch_var = table["garch_var_by_row"][idx]
            restricted_var = table["restricted_var_by_row"][idx]

            ws.cell(idx, 4, float(lag_value))
            ws.cell(idx, 5, float(fitted))
            ws.cell(idx, 6, float(residual))
            ws.cell(idx, 7, float(ewma_var))
            ws.cell(idx, 8, float(math.sqrt(ewma_var)))
            ws.cell(idx, 9, float(garch_var))
            ws.cell(idx, 10, float(math.sqrt(garch_var)))
            ws.cell(idx, 11, float((residual**2) / garch_var))
            ws.cell(idx, 12, float(restricted_var))
            ws.cell(idx, 13, float(math.sqrt(restricted_var)))
            ws.cell(idx, 14, float((residual**2) / restricted_var))

    stats = table["stats"]
    ws["D1"] = stats["intercept"]
    ws["D2"] = stats["phi"]
    ws["F1"] = stats["historical_vol"]
    ws["F2"] = stats["historical_vol"] ** 2
    ws["G3"] = stats["ewma_weight"]
    ws["C4"] = stats["returns_mean"]
    ws["I1"] = stats["omega"]
    ws["I2"] = stats["alpha"]
    ws["I3"] = stats["beta"]
    ws["I4"] = stats["alpha_plus_beta"]
    ws["L2"] = stats["unconditional_variance"]
    ws["N3"] = stats["llf"]
    ws["N4"] = stats["restricted_llf"]
    ws["O1"] = stats["ln_2pi"]
    ws["P4"] = stats["lr_stat"]
    ws["P5"] = stats["chi_sq_95"]

    ws["AA5"] = stats["multiple_r"]
    ws["AA6"] = stats["rsquared"]
    ws["AA7"] = stats["rsquared_adj"]
    ws["AA8"] = stats["standard_error"]
    ws["AA9"] = stats["observations"]

    ws["AA12"] = stats["df_model"]
    ws["AB12"] = stats["regression_ss"]
    ws["AC12"] = stats["regression_ms"]
    ws["AD12"] = stats["f_stat"]
    ws["AE12"] = stats["f_pvalue"]

    ws["AA13"] = stats["df_resid"]
    ws["AB13"] = stats["residual_ss"]
    ws["AC13"] = stats["residual_ms"]

    ws["AA14"] = stats["df_total"]
    ws["AB14"] = stats["total_ss"]

    conf_int = stats["conf_int"]
    params = [stats["intercept"], stats["phi"]]
    coefficient_rows = [18, 19]
    for pos, row in enumerate(coefficient_rows):
        ws.cell(row, 27, params[pos])
        ws.cell(row, 28, stats["std_errs"][pos])
        ws.cell(row, 29, stats["t_stats"][pos])
        ws.cell(row, 30, stats["p_values"][pos])
        ws.cell(row, 31, float(conf_int.iloc[pos, 0]))
        ws.cell(row, 32, float(conf_int.iloc[pos, 1]))
        ws.cell(row, 33, float(conf_int.iloc[pos, 0]))
        ws.cell(row, 34, float(conf_int.iloc[pos, 1]))

    residual_values = [table["predicted_by_row"][row] for row in sorted(table["predicted_by_row"])]
    error_values = [table["residual_by_row"][row] for row in sorted(table["residual_by_row"])]
    for obs_idx, (predicted, error) in enumerate(zip(residual_values, error_values), start=1):
        row = residual_start_row + obs_idx - 1
        ws.cell(row, 26, obs_idx)
        ws.cell(row, 27, float(predicted))
        ws.cell(row, 28, float(error))

    chart_ranges = {
        0: f"'{ws.title}'!$C$7:$C${last_data_row}",
        1: f"'{ws.title}'!$H$8:$H${last_data_row}",
        2: f"'{ws.title}'!$J$11:$J${last_data_row}",
    }
    chart_titles = {
        0: "OVX Returns",
        1: "OVX EWMA Volatility",
        2: "OVX GARCH Volatility",
    }
    for idx, chart in enumerate(ws._charts):
        if idx in chart_ranges and chart.ser:
            chart.ser[0].yVal.numRef.f = chart_ranges[idx]
            chart.title = chart_titles[idx]

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    ovx_df = load_ovx_data()
    returns, reg_df, ols, residuals, garch_fit = fit_models(ovx_df)
    table = build_table(ovx_df, returns, reg_df, ols, residuals, garch_fit)
    output = write_workbook(table)
    print(f"Saved OVX workbook to {output}")
    print(f"Rows written: {len(ovx_df)}")
    print(f"Date range: {ovx_df['Date'].min().date()} to {ovx_df['Date'].max().date()}")
